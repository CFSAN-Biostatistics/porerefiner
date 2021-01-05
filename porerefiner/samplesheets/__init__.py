from porerefiner.protocols.porerefiner.rpc.porerefiner_pb2 import SampleSheet

import csv
from io import TextIOWrapper
from collections import defaultdict

# We should do sample sheet parsing on the client side

# def load_from_csv(file, delimiter=b',') -> SampleSheet:
#     ss = SampleSheet()
#     _, ss.porerefiner_ver, *_ = file.readline().strip().split(delimiter)
#     ss.date.GetCurrentTime()
#     if ss.porerefiner_ver == '1.0.0':
#         _, ss.library_id, *_ = file.readline().strip().split(delimiter)
#         _, ss.sequencing_kit, *_ = file.readline().strip().split(delimiter)
#         delimiter = delimiter.decode()
#         [ss.samples.add(**row) for row in csv.DictReader(TextIOWrapper(file), delimiter=delimiter, dialect='excel')] #this should handle commas in fields
#     elif ss.porerefiner_ver == '1.0.0-fda':
#         _, ss.library_id, *_ = file.readline().strip().split(delimiter)
#         _, ss.sequencing_kit, *_ = file.readline().strip().split(delimiter)
#         delimiter = delimiter.decode()
#         file.readline() #ditch the header
#         for sample_id, accession, barcode_id, organism, extraction_kit, comment, user, *_ in csv.reader(TextIOWrapper(file), delimiter=delimiter, dialect='excel'):
#             ss.samples.add(sample_id=sample_id,
#                            accession=accession,
#                            barcode_id=barcode_id,
#                            organism=organism,
#                            extraction_kit=extraction_kit,
#                            comment=comment,
#                            user=user)
#     elif ss.porerefiner_ver == '1.0.1':
#         _, ss.library_id, *_ = file.readline().strip().split(delimiter)
#         _, ss.sequencing_kit, *_ = file.readline().strip().split(delimiter)
#         # _, ss.barcode_kit, *_ = file.readline().split(delimiter) #TODO
#         # barcode kit takes multiple values
#         _, *barcodes = file.readline().strip().split(delimiter)
#         [ss.barcode_kit.append(barcode) for barcode in barcodes if barcode]
#         delimiter = delimiter.decode()
#         file.readline() # ditch the header
#         for sample_id, accession, barcode_id, organism, extraction_kit, comment, user, *_ in csv.reader(TextIOWrapper(file), delimiter=delimiter, dialect='excel'):
#             ss.samples.add(sample_id=sample_id,
#                            accession=accession,
#                            barcode_id=barcode_id,
#                            organism=organism,
#                            extraction_kit=extraction_kit,
#                            comment=comment,
#                            user=user)
#     else:
#         raise ValueError(f"Sample sheet of version {ss.porerefiner_ver} not supported.")
#     return ss

# def load_from_excel(file) -> SampleSheet:
#     import openpyxl
#     ss = SampleSheet()
#     book = openpyxl.load_workbook(file)
#     rows = (tuple(c.value for c in row) for row in book.worksheets[0].iter_rows())
#     _, ss.porerefiner_ver, *_ = next(rows)
#     if ss.porerefiner_ver == '1.0.0' or ss.porerefiner_ver == '1.0.0-fda':
#         ss.date.GetCurrentTime()
#         _, ss.library_id, *_ = next(rows)
#         _, ss.sequencing_kit, *_ = next(rows)
#         next(rows) # ditch the header
#         for sample_id, accession, barcode_id, organism, extraction_kit, comment, user, *_ in rows:
#             ss.samples.add(sample_id=sample_id,
#                            accession=accession,
#                            barcode_id=str(barcode_id),
#                            organism=organism,
#                            extraction_kit=extraction_kit,
#                            comment=comment,
#                            user=user)
#     elif ss.porerefiner_ver == '1.0.1':
#         ss.date.GetCurrentTime()
#         _, ss.library_id, *_ = next(rows)
#         _, ss.sequencing_kit, *_ = next(rows)
#         _, ss.barcode_kit, *_ = next(rows)
#         next(rows) # ditch the header
#         for sample_id, accession, barcode_id, organism, extraction_kit, comment, user, *_ in rows:
#             ss.samples.add(sample_id=sample_id,
#                            accession=accession,
#                            barcode_id=str(barcode_id),
#                            organism=organism,
#                            extraction_kit=extraction_kit,
#                            comment=comment,
#                            user=user)

#     else:
#         raise ValueError(f"Sample sheet of version {ss.porerefiner_ver} not supported.")

#     return ss



class SnifferFor:

    sniffers = {}

    @classmethod
    def poke(cls, func, format):
        if not hasattr(func, '__formats__'):
            func.__formats__ = set()
        func.__formats__.add(format)
        cls.sniffers[func.__name__] = func
    
    @classmethod
    def csv(cls, sniffer):
        cls.poke(sniffer, 'csv')
        return sniffer

    @classmethod
    def xls(cls, sniffer):
        cls.poke(sniffer, 'xls')
        return sniffer


    class ParserFor:

        def __init__(self):
            self.parsers = {}

        def __getattr__(self, name):
            if name not in SnifferFor.sniffers:
                raise AttributeError(name)
            sniffer = SnifferFor.sniffers[name]
            def parser_decorator(parser):
                self.parsers[sniffer] = parser
                return parser
            return parser_decorator
            

    @classmethod
    def sniff(cls, rows, type='csv'):
        for sniffer in cls.sniffers.values():
            if type in sniffer.__formats__:
                if sniffer(rows):
                    return ParserFor.parsers[sniffer]
                




ParserFor = SnifferFor.ParserFor()


def load_from_excel(file) -> SampleSheet:
    import openpyxl
    book = openpyxl.load_workbook(file)
    rows = (tuple(c.value for c in row) for row in book.worksheets[0].iter_rows())
    return SnifferFor.sniff(rows, type='xls')(rows)

def load_from_csv(file, delimiter=b',') -> SampleSheet:
    rows = [line.split(delimiter) for line in file]
    return SnifferFor.sniff(rows, type='csv')(rows)

